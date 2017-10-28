# -*- coding: utf-8 -*-
"""
Created on Sat Mar 29 00:26:27 2014

@author: Dawid Huczyński
"""
import datetime
import os

import matplotlib.pyplot as plt
import numpy
from openpyxl import Workbook
from openpyxl.chart import LineChart, PieChart, Reference

from config import CONTACTS_FILE, NR_TEL  # Create a config file


class record():
    ''' Not yet implemented, due to nessesity of it.
        Is it easier to work on lists?
    '''

    def __init__(self, data, godzina, typ_pol, kierunek, numer, czas, netto,
                 brutto, pakiet):
        self.data = data
        self.godzina = godzina
        self.typ_pol = typ_pol
        self.kierunek = kierunek
        self.numer = numer
        self.czas = czas
        self.netto = netto
        self.brutto = brutto
        self.pakiet = pakiet

    def get(self, a):
        return self.a


def adressBook(file):
    ''' Imports contacts from exported google contacts.
    file: google.csv file.
    returns: dict {int(tel_number): 'contact_name'}
    '''
    ksiazkaAdresowa = {}
    f = open(file, 'r', encoding='UTF-16')
    t = f.readlines()
    f.close()
    book = [a.strip().split(',') for a in t]
    for contact in book[1:]:
        try:
            if contact[34] != '':
                for number in contact[34].split(':::'):
                    nr = number.replace(' ', '').strip('+')
                    if len(nr) == 9:
                        nr = '48' + nr
                    ksiazkaAdresowa[int(nr)] = contact[0]
        except(KeyError):
            print('Błąd importu dla: %s' % (contact[0]))
        except(IndexError):
            print('Błąd importu dla: %s' % (contact[0]))
    return ksiazkaAdresowa


def scrap_records_TMOBILE():
    ''' Translates records into statistical data.
    records: list with raw txt csv data
    returns: ({'telefoniczne': [], 'SMS': [], 'dane': []}, {}, {})
    '''
    records = []
    files = os.listdir('.')
    for file in files:
        if file.endswith('.txt'):
            f = open(file, 'r', encoding='UTF-8')
            lines = f.readlines()
            telNr = int(lines[1].split('\t')[-1].strip())
            clientId = lines[5].split('\t')[-1].strip()
            records += [x.strip().split('\t') for x in lines[11:]]
    opis = lines[10].strip().split('\t')
    records.sort()
    data, typ_pol, kierunek, numer, czas = 0, 2, 3, 4, 5
    date = records[0][0][4:6]
    month = 0
    polaczenia = {'telefoniczne': [0], 'SMS': [0], 'dane': [0]}
    operatorzy = {}
    rozmowcy = {}
    for wpis in records:
        if wpis[data][4:6] != date:
            date = wpis[data][4:6]
            month += 1
            for k in polaczenia.keys():
                polaczenia[k].append(0)
        if wpis[typ_pol] == 'telefoniczne' and wpis[numer] != 'internet':
            polaczenia[wpis[typ_pol]][
                month] += int(wpis[czas][:2]) * 60 * 60 + \
                int(wpis[czas][3:5]) * 60 + int(wpis[czas][-2:])
            if wpis[kierunek] in operatorzy.keys():
                operatorzy[wpis[kierunek]] += int(wpis[czas][:2]) * \
                    60 * 60 + int(wpis[czas][3:5]) * 60 + int(wpis[czas][-2:])
            else:
                operatorzy[wpis[kierunek]] = int(wpis[czas][:2]) * \
                    60 * 60 + int(wpis[czas][3:5]) * 60 + int(wpis[czas][-2:])
            try:
                rozmowcy[wpis[numer]] += int(wpis[czas][:2]) * 60 * \
                    60 + int(wpis[czas][3:5]) * 60 + int(wpis[czas][-2:])
            except(KeyError):
                rozmowcy[wpis[numer]] = int(wpis[czas][:2]) * 60 * \
                    60 + int(wpis[czas][3:5]) * 60 + int(wpis[czas][-2:])
        elif wpis[typ_pol] == 'SMS':
            polaczenia[wpis[typ_pol]][month] += 1
        elif wpis[typ_pol] == 'telefoniczne' and wpis[numer] == 'internet':
            pass
#            polaczenia[wpis[typ_pol]][month] += int(wpis[czas].strip('kB '))
        if wpis[typ_pol] == 'dane' and wpis[kierunek] in ('internet',
                                                          'Connect internet'):
            polaczenia[wpis[typ_pol]][month] += int(wpis[czas].strip('kB '))
    tel = numpy.array(polaczenia['telefoniczne'])
    print()
    print('Maksymalna ilość pakietów w miesiącu:\t%s' %
          max(polaczenia['dane']))
    print('Maksymalna ilość smsów w miesiącu:\t\t%s' %
          max(polaczenia['SMS']))
    print('Maksymalna ilość wykorzystanych minut:\t%s' %
          max((tel / 60.0).round(0)))
    print()
    print('Średnia pakietów :\t%s' % numpy.mean(polaczenia['dane']))
    print('Średnia smsów:\t%s' % numpy.mean(polaczenia['SMS']).round(2))
    print('Średnia minut:\t%s' % (numpy.mean(tel) / 60).round(2))
    return (polaczenia, operatorzy, rozmowcy, records, telNr, clientId, opis)


def createCharts(sheet, recordsLen, rozmowcyLen, operatorzyLen):
    ''' Creates openpyxl Chart for xls file.
    sheet: openpyxl sheet
    recordsLen: int
    rozmowcyLen: int
    operatorzyLen: int
    returns: chartMB, chartSMS, chartMIN, chartPPL'''
    chartMB = LineChart()
    chartSMS = LineChart()
    chartMIN = LineChart()
    chartPPL = PieChart()
    chartMB.legend = None
    chartSMS.legend = None
    chartMIN.legend = None
    chartMB.title = "MB"
    chartSMS.title = "SMS"
    chartMIN.title = "MIN"
    chartPPL.title = "Ludzie"

    data = Reference(sheet, min_row=7, max_row=recordsLen + 7, min_col=2)
    chartMB.add_data(data, titles_from_data=False)

    data = Reference(sheet, min_row=7, max_row=recordsLen + 7, min_col=3)
    chartSMS.add_data(data, titles_from_data=False)

    data = Reference(sheet, min_row=7, max_row=recordsLen + 7, min_col=4)
    chartMIN.add_data(data, titles_from_data=False)

    dates = Reference(sheet, min_row=7, max_row=recordsLen + 7, min_col=1)
    chartMB.set_categories(dates)
    chartSMS.set_categories(dates)
    chartMIN.set_categories(dates)

    data = Reference(sheet, min_col=6, min_row=5, max_row=5 + rozmowcyLen)
    labels = Reference(sheet, min_col=8, min_row=5, max_row=5 + rozmowcyLen)
    chartPPL.add_data(data)
    chartPPL.set_categories(labels)

    return chartMB, chartSMS, chartMIN, chartPPL


def createArray(records):
    '''Simple beautify function.
    records: list []
    returns: arrray of arrays
    '''
    data = []
    for record in records:
        if record[2] == 'dane' and record[3] == 'internet':
            data.append((record[0], int(record[5].rstrip('kB '))))
    year, month, day = 0, 0, 1
    dayValue = 0
    col = 0
    array = [[]]
    day = 0
    for date, kb in data:
        y, m, d = date[:4], date[4:6], date[6:]
        if year != y or month != m:
            for n in range(31 - len(array[-1])):
                array[-1].append(dayValue)
            array.append([])
            col += 1
            year, month = y, m
            dayValue = 0
        if day != int(d):
            for n in range(int(d) - day):
                array[-1].append(dayValue)
            day = int(d)
            dayValue += kb
        else:
            dayValue += kb
    array = [numpy.array(x) for x in array]
    return array


def monthlyData(records, book):
    '''Writes into openpyxl book monthly data statistisc.
    records: list of lists
    book: openpyxl book
    '''
    data = []
    for record in records:
        if record[2] == 'dane' and record[3] == 'internet':
            data.append((record[0], int(record[5].rstrip('kB '))))
    sheet3 = book.create_sheet('Dane miesiącami')
    year, month, day = 0, 0, 1
    dayValue = 0
    col = 0
    day = 0
    for date, kb in data:
        y, m, d = date[:4], date[4:6], date[6:]
        if year != y or month != m:
            col += 1
            year, month = y, m
            sheet3.cell(row=1, column=col).value = datetime.datetime(
                int(y), int(m), 1)
            sheet3.cell(row=1, column=col).number_format = 'yyyy-mm'
            dayValue = 0
        if day != int(d):
            day = int(d)
            sheet3.cell(row=day + 1, column=col).value = dayValue / 1024
            sheet3.cell(row=day + 1, column=col).number_format = '0.00'
            dayValue += kb
        else:
            dayValue += kb


def writeFile(records, telNr, clientId, polaczenia,
              operatorzy, opis, rozmowcy, ksiazkaAdresowa):
    '''Writes xls file with all the data.
    records: list of lists
    telNr: int
    clientId: int
    polaczenia: dict
    operatorzy: dict
    opis: str
    rozmowcy: dict
    ksiazkaAdresowa: dict
    '''
    book = Workbook(encoding="utf-8")
    sheet = book.active
    sheet.title = "Zestawienie połączeń"
    dates = [x[0] for x in records]
    startDate = min(dates)
    startDate = startDate[:4] + '.' + startDate[4:6] + '.' + startDate[6:]
    endDate = max(dates)
    endDate = endDate[:4] + '.' + endDate[4:6] + '.' + endDate[6:]

    sheet['A1'] = 'Numer telefonu'
    sheet['B1'] = telNr
    sheet['A2'] = 'Numer klienta'
    sheet['B2'] = clientId
    sheet['A3'] = 'Okres rozliczeniowy'
    sheet['B3'] = startDate + '-' + endDate

    sheet.append(['Data', 'MB/msc', 'SMS/msc', 'MIN/msc', '',
                  'Procent rozmów z numerami:', '', '',
                  'Procent rozmów operatorów:', ''])
    for cells in ('B1:J1', 'B2:J2', 'B3:J3', 'I4:J4', 'F4:H4'):
        sheet.merge_cells(cells)
    row = sheet.rows[-1]
    procentRow = sheet.max_row
    maxPakietow = max(polaczenia['dane'])
    if maxPakietow / (1024.0 * 1024) > 1:
        maxPakietow = str(round(maxPakietow / (1024 * 1024), 2)) + 'GB'
    elif maxPakietow / 1024.0 > 1:
        maxPakietow = str(round(maxPakietow / 1024, 2)) + 'MB'
    else:
        maxPakietow = str(maxPakietow) + 'kB'
    srPakietow = numpy.mean(polaczenia['dane'])
    if srPakietow / (1024.0 * 1024) > 1:
        srPakietow = str(round(srPakietow / (1024 * 1024), 2)) + 'GB'
    elif srPakietow / 1024.0 > 1:
        srPakietow = str(round(srPakietow / 1024, 2)) + 'MB'
    else:
        srPakietow = str(srPakietow) + 'kB'
    tel = numpy.array(polaczenia['telefoniczne'])
    sheet.append(['Max', maxPakietow, max(
        polaczenia['SMS']), max((tel / 60.0).round(0))])
    sheet.append(['Średnia', srPakietow, numpy.mean(
        polaczenia['SMS']), (numpy.mean(tel) / 60).round(2)])

    sheet2 = book.create_sheet('Szczegółowa lista połączeń')
    sheet2.append(opis)
    month = 0
    row = sheet2.max_row
    dates = []
    for record in reversed(records):
        dd = datetime.datetime(int(record[0][:4]),
                               int(record[0][4:6]),
                               int(record[0][6:])
                               )
        if dd.month != month:
            month = dd.month
            dates.append(dd)
        if record[4] != '' and record[4][0] not in ['i', 'K', 'e',
                                                    'G', '1', 'C']:
            name = ksiazkaAdresowa.get(int(record[4]))
            if not name:
                name = int(record[4])
        else:
            name = record[4]
        newRecord = [dd] + record[1:4] + [name] + record[5:]
        sheet2.append(newRecord)
        row += 1
        sheet2.cell(row=row, column=1).number_format = 'yyyy-mm-dd'
    monthPack = zip(dates,
                    [round((v / 1024.0), 2)
                     for v in reversed(polaczenia['dane'])],
                    reversed(polaczenia['SMS']),
                    [int(v / 60) for
                     v in reversed(polaczenia['telefoniczne'])])
    for date, MB, SMS, MIN in monthPack:
        sheet.append([date, MB, SMS, MIN])
        sheet.rows[-1][0].number_format = 'yyyy-mm'
    listaRozmowcy = sorted(rozmowcy.items(), key=lambda x: x[1], reverse=True)
    fullRozmowcy = sum(list(rozmowcy.values()))
    rozmowcySum = []
    sheet.column_dimensions['H'].width = 1
    for n, (k, v) in enumerate(listaRozmowcy):
        if int(k) in ksiazkaAdresowa.keys():
            name = ksiazkaAdresowa[int(k)]
        else:
            name = k
        if v / float(fullRozmowcy) < 0.01:
            sheet.cell(row=procentRow + 1 + n, column=6).value = \
                (fullRozmowcy - sum(rozmowcySum)) / float(fullRozmowcy)
            sheet.cell(row=procentRow + 1 + n,
                       column=6).number_format = '0.00%'
            sheet.cell(row=procentRow + 1 + n, column=6 +
                       1).value = fullRozmowcy - sum(rozmowcySum)
            sheet.cell(row=procentRow + 1 + n, column=6 + 2).value = 'Reszta'
            break
        if len(str(name)) > sheet.column_dimensions['H'].width:
            sheet.column_dimensions['H'].width = len(name) + 1
        sheet.cell(row=procentRow + 1 + n, column=6).value = v / \
            float(fullRozmowcy)
        sheet.cell(row=procentRow + 1 + n, column=6).number_format = '0.00%'
        sheet.cell(row=procentRow + 1 + n, column=6 + 1).value = v
        sheet.cell(row=procentRow + 1 + n, column=6 + 2).value = name
        rozmowcySum.append(v)

    full = sum(list(operatorzy.values()))
    operatorzySum = []
    listaOperator = sorted(
        operatorzy.items(), key=lambda x: x[1], reverse=True)
    for n, v in enumerate(listaOperator):
        if v[1] / float(full) < 0.01:
            sheet.cell(row=procentRow + 1 + n, column=9).value = \
                (full - sum(operatorzySum)) / float(full)
            sheet.cell(row=procentRow + 1 + n,
                       column=9).number_format = '0.00%'
            sheet.cell(row=procentRow + 1 + n, column=10).value = "Reszta"
            break
        sheet.cell(row=procentRow + 1 + n, column=9).value = v[1] / float(full)
        sheet.cell(row=procentRow + 1 + n, column=9).number_format = '0.00%'
        sheet.cell(row=procentRow + 1 + n, column=10).value = v[0]
        operatorzySum.append(v[1])

    sheet.append([])
    monthlyData(records, book)
    # Saving
    charts = createCharts(sheet, len(dates), len(
        rozmowcySum), len(operatorzySum))
    sheet.add_chart(charts[0], "E22")
    sheet.add_chart(charts[1], "E37")
    sheet.add_chart(charts[3], sheet.rows[-1][0].coordinate)
    sheet.add_chart(charts[2], "E52")
    charts[0].width = 15
    charts[1].width = 15
    charts[2].width = 15
    charts[3].width = 13
    charts[3].height = 12
    charts[3].legend.position = 'b'
    sheet.column_dimensions['A'].width = 19
    sheet.column_dimensions['J'].width = 16
    sheet.page_setup.orientation = sheet.ORIENTATION_PORTRAIT
    sheet.page_setup.paperSize = sheet.PAPERSIZE_TABLOID
    sheet.page_setup.fitToWidth = 1

    today = datetime.date.today()
    book.save("raport-" + str(telNr) + '-' + str(today.year) +
              '.' + str(today.month) + '.' + str(today.day) + ".xls")


def fitIn(xdata, fit):
    '''Simple fit function.'''
    yfit = []
    for x in xdata:
        y = 0
        for n, v in enumerate(reversed(fit)):
            y += v * (x**n)
        yfit.append(y)
    return yfit


def plots(records, polaczenia):
    '''Writes to plots with statistical data to png files.
    records: list of lists
    polaczenia: dict of lists
    '''
    fig1 = plt.figure()
    fig1.suptitle(str(NR_TEL), size=13)
    ax = fig1.add_subplot(111)
    ydata = createArray(records)[:-1]
    xdata = numpy.array(range(1, 33))
    ax.grid(True)
    plt.xlim(1, 31)
    ylim = 0
    limit = -12
    allY = numpy.array([])
    for y in ydata:
        if max(y) > ylim:
            ylim = max(y)
        ax.fill(xdata, numpy.append(y, [0]) / 1024,
                color=(.2, .7, 0, 1 / len(ydata) + .06))
        ax.plot(xdata, numpy.append(y, [0]) / 1024,
                color=(1, 1, 1, 1 / len(ydata) + .15))
        allY = numpy.append(allY, y)

    ax.plot(xdata[:-1], (sum(ydata[limit:]) / len(ydata[limit:])) /
            1024, color=(1, 0, 0, .4), linestyle='--')

    ax.set_title('Zuzycie MB w ciagu miesiaca', size=11)
    allX = numpy.array(list(xdata)[:-1] * len(ydata[limit:]))
    fit = numpy.polyfit(allX[limit * 31:], allY[limit * 31:], 2)
    fig1.patch.set_facecolor('white')

    yfit = numpy.array(fitIn(xdata, fit)[:-1]) / 1024
    ax.plot(xdata[:-1], yfit, 'b', linestyle='--')
    plt.ylim(0, ylim / 1024 + 100)
    fig1.savefig(str(NR_TEL) + ' pakiety.png')

    xdata = range(len(polaczenia['telefoniczne']))
    teldata = numpy.array(polaczenia['telefoniczne']) / 60
    smsdata = polaczenia['SMS']
    danedata = numpy.array(polaczenia['dane']) / 1024

    f, axrs = plt.subplots(3, sharex=True)
    titles = ['MIN', 'SMS', 'MB']
    for nr, ydata in enumerate([teldata, smsdata, danedata]):
        ax = axrs[nr]
        ax.grid(True)
        ax.plot(xdata, ydata, 'r')
        fit = numpy.polyfit(xdata, ydata, 4)
        xfit = range(len(xdata) + 1)
        yfit = fitIn(xfit, fit)
        ax.plot(xfit, yfit, 'b', linestyle='--')
        ax.set_title(titles[nr], size=11)
    f.suptitle(str(NR_TEL), size=13)
    f.patch.set_facecolor('white')
    f.savefig(str(NR_TEL) + ' wykresy.png')
    plt.show()


def main(NR_TEL):
    path = str(NR_TEL)
    current = os.getcwd()
    print('Folder roboczy: %s' % (os.path.join(current, path)))
    ksiazkaAdresowa = adressBook(CONTACTS_FILE)
    print('Zaimportowano %d adresow.' % (len(ksiazkaAdresowa)))
    os.chdir(os.path.join(os.getcwd(), path))
    scrap = scrap_records_TMOBILE()
    polaczenia, operatorzy, rozmowcy = scrap[:3]
    records, telNr, clientId, opis = scrap[3:]
    os.chdir(current)
    writeFile(records, telNr, clientId, polaczenia, operatorzy, opis, rozmowcy,
              ksiazkaAdresowa)
    plots(records, polaczenia)
    return records, polaczenia, operatorzy, rozmowcy


if __name__ == '__main__':
    # Gets reports data from folder named by telephone number ex \602000111\
    records, polaczenia, operatorzy, rozmowcy = main(NR_TEL)
